"""
数据导出和报表生成模块
"""
import os
import json
import csv
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .database import get_db, Video, Channel
from .utils import format_number, calculate_engagement_rate

class DataExporter:
    """数据导出器"""
    
    def __init__(self):
        self.db = get_db()
    
    def export_channels_report(self, 
                               keyword: str = None,
                               output_format: str = 'excel',
                               output_path: str = None) -> str:
        """导出频道报告"""
        
        session = self.db.get_session()
        
        # 构建查询
        query = session.query(
            Channel.channel_id,
            Channel.title,
            Channel.homepage_url,  # 新增
            Channel.youtube_handle,  # 新增
            Channel.custom_url,
            Channel.country,
            Channel.detected_language,
            Channel.subscriber_count,
            Channel.video_count,
            Channel.view_count,
        )
        
        # 如果指定关键词，添加视频统计
        if keyword:
            video_stats = session.query(
                Video.channel_id,
                func.count(Video.video_id).label('keyword_videos'),
                func.sum(Video.view_count).label('keyword_views'),
                func.avg(Video.view_count).label('avg_views'),
                func.sum(Video.like_count).label('total_likes'),
                func.sum(Video.comment_count).label('total_comments')
            ).filter(
                Video.keyword == keyword
            ).group_by(Video.channel_id).subquery()
            
            query = query.outerjoin(
                video_stats,
                Channel.channel_id == video_stats.c.channel_id
            ).add_columns(
                video_stats.c.keyword_videos,
                video_stats.c.keyword_views,
                video_stats.c.avg_views,
                video_stats.c.total_likes,
                video_stats.c.total_comments
            )
        
        # 执行查询
        results = query.all()
        
        # 转换为DataFrame
        df = self._results_to_dataframe(results, keyword)
        
        # 导出
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"channels_{keyword or 'all'}_{timestamp}"
            output_path = f"./data/{filename}"
        
        if output_format == 'excel':
            output_file = f"{output_path}.xlsx"
            self._export_to_excel(df, output_file, keyword)
        elif output_format == 'csv':
            output_file = f"{output_path}.csv"
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        elif output_format == 'json':
            output_file = f"{output_path}.json"
            df.to_json(output_file, orient='records', force_ascii=False, indent=2)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
        
        session.close()
        
        return output_file
    
    def export_videos_report(self,
                            keyword: str,
                            start_date: datetime = None,
                            end_date: datetime = None,
                            output_format: str = 'excel',
                            output_path: str = None) -> str:
        """导出视频报告"""
        
        session = self.db.get_session()
        
        # 构建查询
        query = session.query(Video).filter(Video.keyword == keyword)
        
        if start_date:
            query = query.filter(Video.published_at >= start_date)
        if end_date:
            query = query.filter(Video.published_at <= end_date)
        
        # 按发布时间排序
        query = query.order_by(Video.published_at.desc())
        
        results = query.all()
        
        # 转换为数据列表
        data = []
        for video in results:
            engagement_rate = calculate_engagement_rate(
                video.view_count,
                video.like_count,
                video.comment_count
            )
            
            data.append({
                'Video ID': video.video_id,
                'Title': video.title,
                'Channel': video.channel_title,
                'Published': video.published_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Views': video.view_count,
                'Likes': video.like_count,
                'Comments': video.comment_count,
                'Engagement Rate': f"{engagement_rate:.2f}%",
                'Duration': video.duration,
                'Duration (seconds)': video.duration_seconds,
                'Is Short': 'Yes' if video.is_short == 1 else 'No',
                'Language': video.language,
                'URL': f"https://youtube.com/watch?v={video.video_id}"
            })
        
        df = pd.DataFrame(data)
        
        # 导出
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"videos_{keyword}_{timestamp}"
            output_path = f"./data/{filename}"
        
        if output_format == 'excel':
            output_file = f"{output_path}.xlsx"
            self._export_videos_to_excel(df, output_file, keyword)
        elif output_format == 'csv':
            output_file = f"{output_path}.csv"
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        elif output_format == 'json':
            output_file = f"{output_path}.json"
            df.to_json(output_file, orient='records', force_ascii=False, indent=2)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
        
        session.close()
        
        return output_file
    
    def generate_summary_report(self, keyword: str = None) -> Dict[str, Any]:
        """生成汇总报告"""
        
        session = self.db.get_session()
        
        report = {
            'generated_at': datetime.now().isoformat(),
            'keyword': keyword or 'All',
            'statistics': {}
        }
        
        if keyword:
            # 特定关键词统计
            video_count = session.query(func.count(Video.video_id)).filter(
                Video.keyword == keyword
            ).scalar()
            
            channel_count = session.query(func.count(func.distinct(Video.channel_id))).filter(
                Video.keyword == keyword
            ).scalar()
            
            total_views = session.query(func.sum(Video.view_count)).filter(
                Video.keyword == keyword
            ).scalar() or 0
            
            avg_views = session.query(func.avg(Video.view_count)).filter(
                Video.keyword == keyword
            ).scalar() or 0
            
            # 时间范围
            date_range = session.query(
                func.min(Video.published_at),
                func.max(Video.published_at)
            ).filter(Video.keyword == keyword).first()
            
            report['statistics'] = {
                'total_videos': video_count,
                'unique_channels': channel_count,
                'total_views': int(total_views),
                'average_views': int(avg_views),
                'date_range': {
                    'start': date_range[0].isoformat() if date_range[0] else None,
                    'end': date_range[1].isoformat() if date_range[1] else None
                }
            }
            
            # Top 频道
            top_channels = session.query(
                Video.channel_id,
                Video.channel_title,
                func.count(Video.video_id).label('video_count'),
                func.sum(Video.view_count).label('total_views')
            ).filter(
                Video.keyword == keyword
            ).group_by(
                Video.channel_id,
                Video.channel_title
            ).order_by(
                func.sum(Video.view_count).desc()
            ).limit(10).all()
            
            report['top_channels'] = [
                {
                    'channel_id': ch[0],
                    'channel_name': ch[1],
                    'video_count': ch[2],
                    'total_views': int(ch[3])
                }
                for ch in top_channels
            ]
            
            # Top 视频
            top_videos = session.query(Video).filter(
                Video.keyword == keyword
            ).order_by(Video.view_count.desc()).limit(10).all()
            
            report['top_videos'] = [
                {
                    'video_id': v.video_id,
                    'title': v.title,
                    'channel': v.channel_title,
                    'views': v.view_count,
                    'published': v.published_at.isoformat()
                }
                for v in top_videos
            ]
            
        else:
            # 全局统计
            total_videos = session.query(func.count(Video.video_id)).scalar()
            total_channels = session.query(func.count(Channel.channel_id)).scalar()
            total_keywords = session.query(func.count(func.distinct(Video.keyword))).scalar()
            
            report['statistics'] = {
                'total_videos': total_videos,
                'total_channels': total_channels,
                'total_keywords': total_keywords
            }
            
            # 关键词统计
            keyword_stats = session.query(
                Video.keyword,
                func.count(Video.video_id).label('count')
            ).group_by(Video.keyword).all()
            
            report['keywords'] = [
                {'keyword': k[0], 'video_count': k[1]}
                for k in keyword_stats
            ]
        
        session.close()
        
        return report
    
    def _results_to_dataframe(self, results: List, keyword: str = None) -> pd.DataFrame:
        """将查询结果转换为DataFrame"""
        
        data = []
        for row in results:
            channel_data = {
                'Channel ID': row[0],
                'Channel Name': row[1],
                'Homepage URL': row[2] if row[2] else f"https://youtube.com/channel/{row[0]}",  # 新增
                'Handle': row[3] or '',  # 新增
                'Custom URL': row[4] or '',
                'Country': row[5] or 'Unknown',
                'Language': row[6] or 'Unknown',
                'Subscribers': format_number(row[7]),
                'Total Videos': row[8],
                'Total Views': format_number(row[9])
            }
            
            if keyword and len(row) > 10:
                # 添加关键词相关统计
                keyword_videos = row[10] or 0
                keyword_views = int(row[11] or 0)
                avg_views = int(row[12] or 0)
                total_likes = int(row[13] or 0)
                total_comments = int(row[14] or 0)
                
                engagement_rate = calculate_engagement_rate(
                    keyword_views,
                    total_likes,
                    total_comments
                )
                
                channel_data.update({
                    f'Videos ({keyword})': keyword_videos,
                    f'Views ({keyword})': format_number(keyword_views),
                    'Avg Views': format_number(avg_views),
                    'Engagement Rate': f"{engagement_rate:.2f}%"
                })
            
            data.append(channel_data)
        
        return pd.DataFrame(data)
    
    def _export_to_excel(self, df: pd.DataFrame, output_file: str, keyword: str = None):
        """导出到Excel文件（带格式）"""
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"Channels - {keyword or 'All'}"
        
        # 添加标题
        ws['A1'] = f"YouTube KOL Report - {keyword or 'All Keywords'}"
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 设置标题样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:J1')
        
        subtitle_font = Font(size=10, italic=True)
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:J2')
        
        # 添加数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 表头样式
                if r_idx == 4:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 数据行样式
                else:
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    # 数字列右对齐
                    if c_idx in [6, 7, 8]:
                        cell.alignment = Alignment(horizontal="right")
        
        # 调整列宽
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                # 跳过合并的单元格
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # 保存文件
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
    
    def _export_videos_to_excel(self, df: pd.DataFrame, output_file: str, keyword: str):
        """导出视频数据到Excel"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Videos - {keyword}"
        
        # 添加标题
        ws['A1'] = f"YouTube Videos Report - {keyword}"
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total Videos: {len(df)}"
        
        # 设置标题样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:K1')
        
        # 添加数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 表头样式
                if r_idx == 5:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # 斑马纹
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        
        # 调整列宽
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                # 跳过合并的单元格
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
