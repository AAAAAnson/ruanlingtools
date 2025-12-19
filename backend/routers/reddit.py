# -*- coding: utf-8 -*-
"""
Reddit keyword search routes

This module handles Reddit-related operations:
- Keyword search across Reddit
- Export results to Excel
"""
import os
import logging
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from typing import Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models.response import ApiResponse
from services.reddit_service import RedditService

router = APIRouter()
logger = logging.getLogger(__name__)


class RedditSearchRequest(BaseModel):
    """Reddit search request model"""
    keyword: str = Field(..., min_length=1, max_length=200, description="Search keyword")
    subreddit: str = Field("all", description="Subreddit name (default: 'all' for site-wide search)")
    limit: int = Field(100, ge=1, le=500, description="Maximum results to return")
    min_comments: int = Field(0, ge=0, description="Minimum comment count filter")
    min_score: int = Field(0, ge=0, description="Minimum score (upvotes) filter")
    time_range_years: int = Field(3, ge=1, le=10, description="Search within last N years")


@router.get("/")
async def reddit_tools_index():
    """
    Get Reddit tools information

    Returns:
        ApiResponse with Reddit tools information
    """
    return ApiResponse.success(
        data={
            "tools": [
                {
                    "id": "keyword-search",
                    "name": "Keyword Search",
                    "description": "Search Reddit posts by keyword and export to Excel",
                    "status": "available"
                }
            ],
            "note": "Reddit features require valid API credentials configured in settings"
        },
        message="Reddit tools available"
    )


@router.post("/search")
async def search_reddit(request: RedditSearchRequest):
    """
    Search Reddit posts by keyword

    This endpoint searches Reddit for posts matching the keyword.
    Filters out promoted posts and returns posts from the last 3 years by default.

    Args:
        request: Reddit search parameters

    Returns:
        ApiResponse containing:
        - keyword: Search keyword used
        - posts: List of matching posts
        - total: Total number of posts found
        - filters: Applied filters
        - timestamp: Search timestamp

    Raises:
        HTTPException: If API credentials are not configured or API error occurs
    """
    try:
        # Initialize Reddit service (will load credentials from settings)
        try:
            reddit_service = RedditService()
        except ValueError as e:
            return ApiResponse.error(
                message="Reddit API credentials not configured. Please configure in Settings.",
                code=503
            )

        # Search Reddit
        results = await reddit_service.search_posts(
            keyword=request.keyword,
            subreddit=request.subreddit,
            limit=request.limit,
            min_comments=request.min_comments,
            min_score=request.min_score,
            time_range_years=request.time_range_years
        )

        if results['total'] == 0:
            return ApiResponse.success(
                data=results,
                message=f"No posts found for keyword: {request.keyword}"
            )

        return ApiResponse.success(
            data=results,
            message=f"Found {results['total']} posts for keyword: {request.keyword}"
        )

    except Exception as e:
        logger.error(f"Reddit search error: {e}", exc_info=True)
        return ApiResponse.error(
            message=f"Failed to search Reddit: {str(e)}",
            code=500
        )


@router.post("/search/export")
async def search_and_export_reddit(request: RedditSearchRequest):
    """
    Search Reddit and export results to Excel

    This endpoint searches Reddit and generates an Excel file with the results.

    Args:
        request: Reddit search parameters

    Returns:
        Excel file download

    Raises:
        HTTPException: If API credentials are not configured or API error occurs
    """
    try:
        # Initialize Reddit service
        try:
            reddit_service = RedditService()
        except ValueError as e:
            raise HTTPException(
                status_code=503,
                detail="Reddit API credentials not configured. Please configure in Settings."
            )

        # Search Reddit
        results = await reddit_service.search_posts(
            keyword=request.keyword,
            subreddit=request.subreddit,
            limit=request.limit,
            min_comments=request.min_comments,
            min_score=request.min_score,
            time_range_years=request.time_range_years
        )

        if results['total'] == 0:
            raise HTTPException(
                status_code=404,
                detail=f"No posts found for keyword: {request.keyword}"
            )

        # Format for export
        export_data = reddit_service.format_for_export(results)

        # Create Excel file
        excel_path = await create_excel_file(
            export_data,
            keyword=request.keyword,
            filters=results['filters']
        )

        # Return file
        filename = f"reddit_search_{request.keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return FileResponse(
            path=excel_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Reddit export error: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export Reddit search: {str(e)}"
        )


async def create_excel_file(data: list, keyword: str, filters: dict) -> str:
    """
    Create Excel file from search results

    Args:
        data: List of post dictionaries
        keyword: Search keyword
        filters: Applied filters

    Returns:
        Path to created Excel file
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reddit Search Results"

    # Add title
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"Reddit Search Results - Keyword: {keyword}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add filter info
    ws.merge_cells('A2:N2')
    filter_cell = ws['A2']
    filter_info = f"Filters: Time Range={filters.get('time_range_years')}y, Min Comments={filters.get('min_comments')}, Min Score={filters.get('min_score')}, Exclude Promoted={filters.get('exclude_promoted')}"
    filter_cell.value = filter_info
    filter_cell.font = Font(size=10, italic=True)
    filter_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add timestamp
    ws.merge_cells('A3:N3')
    timestamp_cell = ws['A3']
    timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    timestamp_cell.font = Font(size=9, italic=True)
    timestamp_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add headers (row 5)
    if data:
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for row_num, row_data in enumerate(data, 6):
            for col_num, (key, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column].width = adjusted_width

        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[5].height = 25

    # Ensure outputs directory exists
    outputs_dir = os.path.join(os.path.dirname(__file__), '..', 'outputs')
    os.makedirs(outputs_dir, exist_ok=True)

    # Save file
    filename = f"reddit_search_{keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(outputs_dir, filename)
    wb.save(filepath)

    logger.info(f"Excel file created: {filepath}")
    return filepath


@router.get("/config")
async def get_reddit_config():
    """
    Get Reddit API configuration status

    Returns:
        ApiResponse with configuration status
    """
    # Check if credentials are available
    from services.settings_service import get_settings_service

    try:
        settings_service = get_settings_service()
        reddit_config = settings_service.get_reddit_config()

        configured = bool(
            reddit_config and
            reddit_config.get('client_id') and
            reddit_config.get('client_secret') and
            reddit_config.get('user_agent')
        )

        return ApiResponse.success(
            data={
                "api_configured": configured,
                "features_available": configured
            },
            message="Reddit configuration status"
        )
    except Exception as e:
        return ApiResponse.success(
            data={
                "api_configured": False,
                "features_available": False
            },
            message="Reddit not configured"
        )
